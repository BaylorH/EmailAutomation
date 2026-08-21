#!/usr/bin/env python3
"""Is it safe to launch this campaign right now? Run BEFORE every launch.

This is the check that replaces asking permission. The standing test identities are
pre-authorized, so the question is never "may I" -- it is "does this specific thing
reach anyone it must not, and is anything else already in flight". That is a question
with an answer, and this answers it.

    cd <repo>
    set -a; source .env; set +a
    GOOGLE_APPLICATION_CREDENTIALS=$PWD/service-account.json \
        python3 scripts/preflight_campaign.py campaign.xlsx --message body.txt

Exit 0 = safe to launch. Exit 1 = STOP, do not launch, fix what it names.

WHAT IT CHECKS, and why each one is here rather than trusted:

1. EVERY address anywhere in the workbook -- not just the Email column. A stray address
   in a comments cell is still an address the product can pick up, and "I only looked at
   the column I meant to fill" is exactly how one leaks.
2. EVERY address in the outgoing message body, if given. A referral or a signature line
   written into the copy is a live address; this was a real near-miss when scripting a
   wrong-contact test.
3. NO OTHER CAMPAIGN is live anywhere, across ALL users -- not just the one the dashboard
   happens to be showing. Other mailboxes hold real third-party correspondence.
4. NOTHING IS ALREADY QUEUED OR ARMED, by delegating to the send-exposure audit so there
   is exactly one allow-list in the repo and it cannot drift.

It reads Firestore and a local file. It sends nothing and writes nothing.
"""
import argparse
import importlib.util
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))

# Single source of truth for the allow-list: the audit's own predicate. Importing it
# rather than restating it is deliberate -- two copies of an allow-list is one copy that
# is quietly wrong.
_spec = importlib.util.spec_from_file_location(
    "audit_send_exposure", os.path.join(HERE, "audit_send_exposure.py")
)
_audit = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_audit)
allowed = _audit.allowed

TERMINAL_CLIENT = _audit.TERMINAL_CLIENT

# Deliberately greedy: it is better to flag a string that merely looks like an address
# than to miss one. Everything it finds is checked against the allow-list anyway.
ADDRESS_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def addresses_in_workbook(path):
    """Every address-shaped string in EVERY cell of EVERY sheet."""
    from openpyxl import load_workbook

    found = {}
    book = load_workbook(path, data_only=True)
    for sheet in book.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                for hit in ADDRESS_RE.findall(str(cell.value)):
                    found.setdefault(hit.lower(), []).append(
                        f"{sheet.title}!{cell.coordinate}"
                    )
    return found


def addresses_in_text(text):
    found = {}
    for hit in ADDRESS_RE.findall(text or ""):
        found.setdefault(hit.lower(), []).append("message body")
    return found


def other_live_campaigns(fs):
    """Live campaigns across ALL users. Returns a list of human-readable strings."""
    live = []
    for user in fs.collection("users").stream():
        base = fs.collection("users").document(user.id)
        for client in base.collection("clients").stream():
            data = client.to_dict() or {}
            status = str(data.get("status") or "").strip().lower()
            if status in TERMINAL_CLIENT:
                continue
            live.append(f"{data.get('name', '?')!r} [{user.id[:10]}] status={status}")
    return live


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", nargs="?", help="the .xlsx about to be uploaded")
    ap.add_argument("--message", help="file containing the outgoing message body")
    ap.add_argument(
        "--expect-live",
        nargs="*",
        default=[],
        help="client ids already expected to be live (a campaign mid-test)",
    )
    args = ap.parse_args()

    problems, notes = [], []

    # ---- 1 & 2: nothing addressed outside the owned accounts ------------------
    seen = {}
    if args.workbook:
        seen.update(addresses_in_workbook(args.workbook))
    if args.message:
        with open(args.message, encoding="utf-8") as handle:
            for addr, where in addresses_in_text(handle.read()).items():
                seen.setdefault(addr, []).extend(where)

    foreign = {a: w for a, w in seen.items() if not allowed(a)}
    if foreign:
        for addr, where in sorted(foreign.items()):
            problems.append(f"FOREIGN ADDRESS {addr} at {', '.join(sorted(set(where)))}")
    elif seen:
        notes.append(f"{len(seen)} address(es) found, all self-owned")

    # ---- 3: nothing else in flight -------------------------------------------
    from google.cloud import firestore

    fs = firestore.Client()
    expected = set(args.expect_live or ())
    for entry in other_live_campaigns(fs):
        if any(cid in entry for cid in expected):
            notes.append(f"live (expected): {entry}")
        else:
            problems.append(f"ANOTHER CAMPAIGN IS LIVE: {entry}")

    # ---- 4: nothing queued or armed ------------------------------------------
    audit_problems, audit_notes = _audit.audit(fs, expected)
    problems.extend(audit_problems)
    notes.extend(audit_notes)

    print("=== CAMPAIGN PREFLIGHT ===")
    for note in notes:
        print("  note:", note)
    if problems:
        print("\nDO NOT LAUNCH:")
        for problem in problems:
            print("  !!", problem)
        return 1
    print("  clear: nothing addressed outside the owned accounts, nothing else in flight")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
