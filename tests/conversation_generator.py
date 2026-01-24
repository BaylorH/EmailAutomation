#!/usr/bin/env python3
"""
Conversation Generator
======================
Programmatically generates conversation files to cover all possible
real-world scenarios. This helps map out everything that might happen
in production.

Usage:
    python3 tests/conversation_generator.py --generate-all
    python3 tests/conversation_generator.py --list-scenarios
    python3 tests/conversation_generator.py --property "1 Kuhlke Dr" --scenario complete_info

Scenarios covered:
    - Broker response types (complete, partial, vague, hostile)
    - Event scenarios (call requested, property unavailable, new property)
    - Edge cases (forwarding, out of office, multiple contacts)
    - Number format variations (words, commas, decimals)
    - Question types (scheduling, client info, pricing)
"""

import os
import sys
import json
from pathlib import Path
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ============================================================================
# SCENARIO DEFINITIONS
# ============================================================================

@dataclass
class Scenario:
    """A test scenario definition."""
    name: str
    description: str
    category: str  # response_type, event, edge_case, format
    inbound_template: str  # Template for broker's reply
    expected_updates: List[Dict]
    expected_events: List[str]
    forbidden_updates: List[str]
    notes: str = ""


# Response type scenarios
RESPONSE_SCENARIOS = [
    Scenario(
        name="complete_info",
        description="Broker provides all property specs in one reply",
        category="response_type",
        inbound_template="""Hi,

Happy to share the details for {property}:

- Total SF: {sf}
- Asking: ${rent}/SF/yr NNN
- NNN/OpEx: ${opex}/SF/yr
- {driveins} drive-in doors
- {docks} dock-high doors
- {clearht}' clear height
- {power} power

Let me know if you have questions!

{contact}""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"},
            {"column": "Ops Ex /SF"},
            {"column": "Drive Ins"},
            {"column": " Docks"},
            {"column": "Ceiling Ht"},
            {"column": "Power"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"]
    ),

    Scenario(
        name="partial_info",
        description="Broker provides only some specs, AI should request missing",
        category="response_type",
        inbound_template="""Hi,

Here's what I have for {property}:
- SF: {sf}
- Asking {rent}/SF NNN

I'll get back to you on the other specs.

{contact}""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should request: NNN/OpEx, doors, clear height, power"
    ),

    Scenario(
        name="vague_response",
        description="Broker gives non-specific answer with no actual data",
        category="response_type",
        inbound_template="""Hi,

Thanks for reaching out about {property}. It's a nice space in a great location with good access.

Let me know if you want to schedule a tour.

{contact}""",
        expected_updates=[],
        expected_events=["needs_user_input"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should re-request specific specs"
    ),

    Scenario(
        name="terse_response",
        description="Broker gives very short, minimal reply",
        category="response_type",
        inbound_template="""15000 sf, $7.50 nnn""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"]
    )
]

# Event scenarios
EVENT_SCENARIOS = [
    Scenario(
        name="call_requested_with_phone",
        description="Broker wants to discuss by phone, provides number",
        category="event",
        inbound_template="""Hi,

I'd prefer to discuss {property} over the phone. Can you call me at {phone}?

Thanks,
{contact}""",
        expected_updates=[],
        expected_events=["call_requested"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="Phone number provided - no response email needed"
    ),

    Scenario(
        name="call_requested_no_phone",
        description="Broker wants a call but doesn't provide number",
        category="event",
        inbound_template="""Hi,

Would you be able to give me a call to discuss {property}? Easier to explain in conversation.

{contact}""",
        expected_updates=[],
        expected_events=["call_requested"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="No phone number - AI may ask for phone number"
    ),

    Scenario(
        name="property_unavailable",
        description="Broker says property is no longer available",
        category="event",
        inbound_template="""Hi,

Unfortunately {property} is no longer available - we just signed a lease last week.

I can let you know if anything else comes up in the area.

{contact}""",
        expected_updates=[],
        expected_events=["property_unavailable"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should ask about alternatives"
    ),

    Scenario(
        name="unavailable_with_alternative",
        description="Property unavailable but broker suggests alternative",
        category="event",
        inbound_template="""Hi,

{property} was just leased. However, I have another space at {alt_property} in {alt_city} that might work:

- {alt_sf} SF
- ${alt_rent}/SF/yr NNN

Want me to send over more info?

{contact}""",
        expected_updates=[],
        expected_events=["property_unavailable", "new_property"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should handle unavailable + propose new property"
    ),

    Scenario(
        name="new_property_suggestion",
        description="Broker proactively suggests another property",
        category="event",
        inbound_template="""Hi,

Here are the specs for {property}:
- {sf} SF
- ${rent}/SF NNN

Also, I wanted to mention we have another property at {alt_property} that might work for your client - {alt_sf} SF with similar specs. Let me know if interested.

{contact}""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"}
        ],
        expected_events=["new_property"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"]
    ),

    Scenario(
        name="contact_optout",
        description="Broker asks to be removed from contact list",
        category="event",
        inbound_template="""Please remove me from your mailing list. I don't work with tenant reps.

Do not contact me again.""",
        expected_updates=[],
        expected_events=["contact_optout"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="No auto-response should be sent"
    )
]

# Edge case scenarios
EDGE_SCENARIOS = [
    Scenario(
        name="forward_to_colleague",
        description="Broker forwards to colleague - should NOT update contact info",
        category="edge_case",
        inbound_template="""Hi,

I'm forwarding this to my colleague {new_contact} who handles {property}. She'll reach out to you directly.

{contact}""",
        expected_updates=[],
        expected_events=["wrong_contact"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should NOT update contact to the colleague"
    ),

    Scenario(
        name="out_of_office",
        description="Broker is out of office with auto-reply",
        category="edge_case",
        inbound_template="""I am currently out of the office with limited access to email.

I will return on Monday, January 20th.

For urgent matters, please contact {backup_contact} at {backup_email}.

{contact}""",
        expected_updates=[],
        expected_events=["needs_user_input"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should wait or escalate, not auto-respond"
    ),

    Scenario(
        name="flyer_link_only",
        description="Broker sends only a flyer link",
        category="edge_case",
        inbound_template="""Here's the flyer for {property}: {flyer_url}

Let me know if you have questions.

{contact}""",
        expected_updates=[
            {"column": "Flyer / Link"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should extract link AND request actual specs"
    ),

    Scenario(
        name="question_about_client",
        description="Broker asks about our client's requirements",
        category="edge_case",
        inbound_template="""Hi,

Before I send specs, what size is your client looking for? And what's their timeline?

{contact}""",
        expected_updates=[],
        expected_events=["needs_user_input"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should escalate - user needs to provide client info"
    ),

    Scenario(
        name="tour_offer",
        description="Broker offers to schedule a tour",
        category="edge_case",
        inbound_template="""Hi,

Here are the specs for {property}:
- {sf} SF, ${rent}/SF NNN
- {docks} dock doors, {driveins} drive-in
- {clearht}' clear

Would your client like to tour the space? I'm available this week.

{contact}""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"},
            {"column": " Docks"},
            {"column": "Drive Ins"},
            {"column": "Ceiling Ht"}
        ],
        expected_events=["needs_user_input"],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should extract data AND escalate tour question"
    ),

    Scenario(
        name="pricing_negotiable",
        description="Broker indicates pricing is negotiable",
        category="edge_case",
        inbound_template="""Hi,

For {property}:
- {sf} SF
- Asking ${rent}/SF but the owner is motivated so pricing is negotiable
- NNN around ${opex}

Happy to discuss further.

{contact}""",
        expected_updates=[
            {"column": "Total SF"},
            {"column": "Rent/SF /Yr"},
            {"column": "Ops Ex /SF"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should note 'negotiable' in response or comments"
    )
]

# Number format variation scenarios
FORMAT_SCENARIOS = [
    Scenario(
        name="numbers_with_words",
        description="Broker uses word numbers - AI should normalize",
        category="format",
        inbound_template="""Hi,

{property} specs:
- eighteen thousand five hundred square feet
- seven fifty per foot triple net
- about two fifteen in NNN
- three dock doors, one drive-in
- twenty-four foot clear
- two hundred amp three-phase

{contact}""",
        expected_updates=[
            {"column": "Total SF", "value": "18500"},
            {"column": "Rent/SF /Yr", "value": "7.50"},
            {"column": "Ops Ex /SF", "value": "2.15"},
            {"column": " Docks", "value": "3"},
            {"column": "Drive Ins", "value": "1"},
            {"column": "Ceiling Ht", "value": "24"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"],
        notes="AI should normalize 'eighteen thousand' to '18500' etc"
    ),

    Scenario(
        name="numbers_with_mixed_formats",
        description="Broker uses various number formats",
        category="format",
        inbound_template="""Hi,

Details on {property}:
- Size: 22,500 SF
- Rent: $6.75/sf/yr NNN
- NNN: ~$2.00/SF/YR
- Doors: 2 DI, 4 DH
- Clear: 26'
- Power: 400A 3ph

{contact}""",
        expected_updates=[
            {"column": "Total SF", "value": "22500"},
            {"column": "Rent/SF /Yr", "value": "6.75"},
            {"column": "Ops Ex /SF", "value": "2.00"},
            {"column": "Drive Ins", "value": "2"},
            {"column": " Docks", "value": "4"},
            {"column": "Ceiling Ht", "value": "26"}
        ],
        expected_events=[],
        forbidden_updates=["Leasing Contact", "Leasing Company", "Email"]
    )
]

ALL_SCENARIOS = RESPONSE_SCENARIOS + EVENT_SCENARIOS + EDGE_SCENARIOS + FORMAT_SCENARIOS


# ============================================================================
# PROPERTY DATA
# ============================================================================

# Sample property values for filling templates
SAMPLE_VALUES = {
    "sf": "15000",
    "rent": "7.50",
    "opex": "2.15",
    "driveins": "2",
    "docks": "4",
    "clearht": "24",
    "power": "200 amp 3-phase",
    "phone": "(706) 555-1234",
    "alt_property": "456 Commerce Blvd",
    "alt_city": "Augusta",
    "alt_sf": "12000",
    "alt_rent": "6.50",
    "new_contact": "Sarah Johnson",
    "backup_contact": "Mike Smith",
    "backup_email": "mike@broker.com",
    "flyer_url": "https://example.com/flyers/property.pdf"
}


# ============================================================================
# GENERATOR
# ============================================================================

def load_properties() -> Dict:
    """Load properties from Scrub Excel file."""
    from openpyxl import load_workbook

    scrub_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "Scrub Augusta GA.xlsx"
    )

    if not os.path.exists(scrub_path):
        return {}

    wb = load_workbook(scrub_path)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]

    properties = {}
    for row_num in range(3, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_num]]
        if not row_values or not row_values[0]:
            continue

        address = str(row_values[0]).strip()
        city = str(row_values[1] or "").strip()
        contact = str(row_values[4] or "").strip()

        properties[address] = {
            "city": city,
            "contact": contact
        }

    return properties


def generate_conversation(property_address: str, property_data: Dict, scenario: Scenario) -> Dict:
    """Generate a conversation file for a property and scenario."""
    contact = property_data.get("contact", "Broker")
    city = property_data.get("city", "")

    # Build template values
    values = {
        **SAMPLE_VALUES,
        "property": property_address,
        "contact": contact,
        "city": city
    }

    # Fill template
    inbound_content = scenario.inbound_template.format(**values)

    # Build outbound message
    outbound = f"Hi {contact}, I'm reaching out about {property_address}"
    if city:
        outbound += f" in {city}"
    outbound += ". Could you share the property details including SF, asking rent, NNN, loading doors, clear height, and power?"

    conversation = {
        "property": property_address,
        "city": city,
        "description": scenario.description,
        "messages": [
            {"direction": "outbound", "content": outbound},
            {"direction": "inbound", "content": inbound_content}
        ],
        "expected_updates": scenario.expected_updates,
        "expected_events": scenario.expected_events,
        "forbidden_updates": scenario.forbidden_updates,
        "notes": scenario.notes
    }

    return conversation


def save_conversation(conversation: Dict, scenario: Scenario, output_dir: Path):
    """Save a conversation to a JSON file."""
    # Create category subdirectory
    category_dir = output_dir / scenario.category
    category_dir.mkdir(exist_ok=True)

    # Filename from property and scenario
    prop_name = conversation["property"].lower().replace(" ", "_").replace(",", "")
    filename = f"{prop_name}_{scenario.name}.json"

    filepath = category_dir / filename
    with open(filepath, "w") as f:
        json.dump(conversation, f, indent=2)

    return filepath


def generate_all_conversations(output_dir: Path = None):
    """Generate conversations for all properties and scenarios."""
    if output_dir is None:
        output_dir = Path(__file__).parent / "conversations" / "generated"

    output_dir.mkdir(parents=True, exist_ok=True)

    properties = load_properties()
    if not properties:
        print("No properties found in Scrub file")
        return

    print(f"Generating conversations for {len(properties)} properties x {len(ALL_SCENARIOS)} scenarios...")

    generated = []
    for prop_addr, prop_data in properties.items():
        for scenario in ALL_SCENARIOS:
            conv = generate_conversation(prop_addr, prop_data, scenario)
            filepath = save_conversation(conv, scenario, output_dir)
            generated.append(str(filepath))

    # Create index file
    index = {
        "generated_at": str(Path(__file__).parent / "conversations" / "generated"),
        "property_count": len(properties),
        "scenario_count": len(ALL_SCENARIOS),
        "total_conversations": len(generated),
        "categories": {
            "response_type": len(RESPONSE_SCENARIOS),
            "event": len(EVENT_SCENARIOS),
            "edge_case": len(EDGE_SCENARIOS),
            "format": len(FORMAT_SCENARIOS)
        },
        "files": generated
    }

    with open(output_dir / "index.json", "w") as f:
        json.dump(index, f, indent=2)

    print(f"Generated {len(generated)} conversation files in {output_dir}")
    return generated


def list_scenarios():
    """List all available scenarios."""
    print("\nAvailable Scenarios:")
    print("=" * 60)

    for category, scenarios in [
        ("Response Types", RESPONSE_SCENARIOS),
        ("Events", EVENT_SCENARIOS),
        ("Edge Cases", EDGE_SCENARIOS),
        ("Format Variations", FORMAT_SCENARIOS)
    ]:
        print(f"\n{category}:")
        for s in scenarios:
            print(f"  {s.name:30} - {s.description}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate test conversations")
    parser.add_argument("--generate-all", action="store_true", help="Generate all conversations")
    parser.add_argument("--list-scenarios", action="store_true", help="List available scenarios")
    parser.add_argument("--property", help="Generate for specific property")
    parser.add_argument("--scenario", help="Generate specific scenario")
    parser.add_argument("--output", help="Output directory")
    args = parser.parse_args()

    if args.list_scenarios:
        list_scenarios()
        return

    if args.generate_all:
        output_dir = Path(args.output) if args.output else None
        generate_all_conversations(output_dir)
        return

    if args.property and args.scenario:
        properties = load_properties()
        if args.property not in properties:
            print(f"Property not found: {args.property}")
            print("Available:", list(properties.keys()))
            return

        scenario = next((s for s in ALL_SCENARIOS if s.name == args.scenario), None)
        if not scenario:
            print(f"Scenario not found: {args.scenario}")
            list_scenarios()
            return

        conv = generate_conversation(args.property, properties[args.property], scenario)
        print(json.dumps(conv, indent=2))
        return

    parser.print_help()


if __name__ == "__main__":
    main()
