#!/usr/bin/env python3
"""
End-to-End Integration Test Framework
======================================
Tests the FULL production pipeline by:
1. Loading property data from the Scrub Excel file
2. Using conversation files that define broker replies
3. Running the ACTUAL production code paths
4. Verifying sheet updates, notifications, and response emails

This ensures tests are a 1:1 reflection of what happens in production.

Results Output:
    When --save is used, results are saved to tests/results/run_YYYYMMDD_HHMMSS/
    Each result file contains: input data, conversation, AI output, sheet state,
    notifications, and validation results.

Usage:
    python tests/e2e_test.py                    # Run all E2E tests
    python tests/e2e_test.py --save             # Run and save results to files
    python tests/e2e_test.py -p "699 Industrial" # Run specific property
    python tests/e2e_test.py --list             # List available conversations
    python tests/e2e_test.py --list-runs        # List previous test runs
    python tests/e2e_test.py --compare run1 run2 # Compare two runs
"""

import os
import sys
import json
import copy
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

# Results management - import after path setup below

# ============================================================================
# ENVIRONMENT SETUP
# ============================================================================

# Load .env file
env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ[key] = value

# Check for OpenAI API key
if not os.getenv("OPENAI_API_KEY"):
    print("OPENAI_API_KEY environment variable not set")
    sys.exit(1)

# Set dummy Azure env vars if not present
for var in ["AZURE_API_APP_ID", "AZURE_API_CLIENT_SECRET", "FIREBASE_API_KEY"]:
    if not os.getenv(var):
        os.environ[var] = f"test-{var.lower()}"

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Results management - import now that path is set up
from tests.results_manager import (
    create_run_directory,
    create_manifest,
    save_result,
    save_summary,
    list_runs,
    load_run,
    compare_runs
)

# ============================================================================
# MOCK INFRASTRUCTURE
# ============================================================================

# Captured outputs from production code
CAPTURED_SHEET_UPDATES = []
CAPTURED_NOTIFICATIONS = []
CAPTURED_OUTBOX_EMAILS = []

def reset_captures():
    """Reset all captured outputs between tests."""
    global CAPTURED_SHEET_UPDATES, CAPTURED_NOTIFICATIONS, CAPTURED_OUTBOX_EMAILS
    CAPTURED_SHEET_UPDATES = []
    CAPTURED_NOTIFICATIONS = []
    CAPTURED_OUTBOX_EMAILS = []

# Mock Firestore
from unittest.mock import MagicMock, patch
import sys as _sys

mock_firestore = MagicMock()
mock_firestore.Client = MagicMock(return_value=MagicMock())
mock_firestore.SERVER_TIMESTAMP = "SERVER_TIMESTAMP"
_sys.modules['google.cloud.firestore'] = mock_firestore
_sys.modules['google.cloud'] = MagicMock()
_sys.modules['google.oauth2.credentials'] = MagicMock()
_sys.modules['google.auth.transport.requests'] = MagicMock()
_sys.modules['googleapiclient.discovery'] = MagicMock()

# Now import production code
from email_automation.ai_processing import propose_sheet_updates
from email_automation.app_config import REQUIRED_FIELDS_FOR_CLOSE

# ============================================================================
# SCRUB FILE LOADER
# ============================================================================

def load_scrub_file(filepath: str = None) -> Dict[str, Dict]:
    """
    Load the Scrub Excel file and return property data.
    Returns: {property_address: {row, city, contact, email, data, header}}
    """
    if filepath is None:
        filepath = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "Scrub Augusta GA.xlsx"
        )

    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    ws = wb.active

    # Get headers from row 2
    headers = [cell.value for cell in ws[2]]

    properties = {}
    for row_num in range(3, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_num]]

        # Skip empty rows
        if not row_values or not row_values[0]:
            continue

        address = str(row_values[0]).strip()
        city = str(row_values[1] or "").strip()
        contact = str(row_values[4] or "").strip()
        email = str(row_values[5] or "").strip()

        # Normalize row values (convert None to empty string)
        data = [str(v) if v is not None else "" for v in row_values]

        properties[address] = {
            "row": row_num,
            "city": city,
            "contact": contact,
            "email": email,
            "data": data,
            "header": headers
        }

    return properties

# ============================================================================
# CONVERSATION FILE LOADER
# ============================================================================

def get_conversations_dir() -> Path:
    """Get the conversations directory path."""
    return Path(__file__).parent / "conversations"

def load_conversation(property_address: str, subdir: str = None) -> Optional[Dict]:
    """
    Load a conversation file for a property.
    Returns: {messages: [...], expected_updates: [...], expected_events: [...], ...}

    Args:
        property_address: The property address to load conversation for
        subdir: Optional subdirectory (e.g., "edge_cases")
    """
    conv_dir = get_conversations_dir()
    if subdir:
        conv_dir = conv_dir / subdir

    # Normalize address to filename
    filename = property_address.lower().replace(" ", "_").replace(",", "")

    # Try different extensions
    for ext in [".json", ".jsonc"]:
        filepath = conv_dir / f"{filename}{ext}"
        if filepath.exists():
            with open(filepath) as f:
                return json.load(f)

    return None


def load_all_edge_case_conversations() -> List[Tuple[str, Dict]]:
    """
    Load all edge case conversation files.
    Returns: [(filename, conversation_dict), ...]
    """
    edge_dir = get_conversations_dir() / "edge_cases"
    if not edge_dir.exists():
        return []

    conversations = []
    for f in sorted(edge_dir.glob("*.json")):
        with open(f) as fp:
            conv = json.load(fp)
            conversations.append((f.stem, conv))

    return conversations

def list_available_conversations() -> List[str]:
    """List all available conversation files."""
    conv_dir = get_conversations_dir()
    if not conv_dir.exists():
        return []

    conversations = []
    for f in conv_dir.glob("*.json"):
        conversations.append(f.stem.replace("_", " ").title())

    return sorted(conversations)


def load_generated_conversations(category: str = "all") -> List[Tuple[str, Dict]]:
    """
    Load generated conversation files from tests/conversations/generated/.

    Args:
        category: 'response_type', 'event', 'edge_case', 'format', or 'all'

    Returns: [(label, conversation_dict), ...]
    """
    gen_dir = get_conversations_dir() / "generated"
    if not gen_dir.exists():
        return []

    conversations = []
    categories = ["response_type", "event", "edge_case", "format"] if category == "all" else [category]

    for cat in categories:
        cat_dir = gen_dir / cat
        if not cat_dir.exists():
            continue

        for f in sorted(cat_dir.glob("*.json")):
            with open(f) as fp:
                conv = json.load(fp)
                label = f"[GEN:{cat}] {f.stem}"
                conversations.append((label, conv))

    return conversations

# ============================================================================
# TEST EXECUTION
# ============================================================================

@dataclass
class E2ETestResult:
    """Result of an E2E test."""
    property_address: str
    passed: bool = False

    # AI outputs
    ai_updates: List[Dict] = field(default_factory=list)
    ai_events: List[Dict] = field(default_factory=list)
    ai_response: str = ""
    ai_notes: str = ""

    # Simulated sheet state
    sheet_before: List[str] = field(default_factory=list)
    sheet_after: List[str] = field(default_factory=list)

    # Notifications
    notifications: List[Dict] = field(default_factory=list)

    # Validation
    issues: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def run_e2e_test(property_address: str, property_data: Dict, conversation: Dict) -> E2ETestResult:
    """
    Run a full E2E test for a single property.

    This calls the ACTUAL production propose_sheet_updates() function,
    then verifies the outputs against expected values.
    """
    result = E2ETestResult(property_address=property_address)
    reset_captures()

    # Build conversation payload (same format as messaging.py build_conversation_payload)
    messages = conversation.get("messages", [])
    conv_payload = []

    for i, msg in enumerate(messages):
        conv_payload.append({
            "direction": msg["direction"],
            "from": property_data["email"] if msg["direction"] == "inbound" else "jill@company.com",
            "to": ["jill@company.com"] if msg["direction"] == "inbound" else [property_data["email"]],
            "subject": f"{property_address}, {property_data['city']}",
            "timestamp": f"2024-01-15T{10+i}:00:00Z",
            "preview": msg["content"][:200],
            "content": msg["content"]
        })

    # Store sheet state before
    result.sheet_before = list(property_data["data"])

    # Call PRODUCTION code
    try:
        proposal = propose_sheet_updates(
            uid="e2e-test-user",
            client_id="e2e-test-client",
            email=property_data["email"],
            sheet_id="e2e-test-sheet",
            header=property_data["header"],
            rownum=property_data["row"],
            rowvals=property_data["data"],
            thread_id=f"e2e-thread-{property_address.lower().replace(' ', '-')}",
            contact_name=property_data["contact"],
            conversation=conv_payload,
            dry_run=True
        )

        if proposal:
            result.ai_updates = proposal.get("updates", [])
            result.ai_events = proposal.get("events", [])
            result.ai_response = proposal.get("response_email", "")
            result.ai_notes = proposal.get("notes", "")

    except Exception as e:
        result.issues.append(f"Production code error: {str(e)}")
        return result

    # Apply updates to get sheet_after state
    result.sheet_after = list(property_data["data"])
    header_lower = {h.lower().strip() if h else "": i for i, h in enumerate(property_data["header"])}

    for update in result.ai_updates:
        col_name = update.get("column", "").lower().strip()
        if col_name in header_lower:
            idx = header_lower[col_name]
            if idx < len(result.sheet_after):
                result.sheet_after[idx] = update.get("value", "")

    # Derive notifications that would fire
    for update in result.ai_updates:
        result.notifications.append({
            "kind": "sheet_update",
            "column": update.get("column"),
            "value": update.get("value")
        })

    for event in result.ai_events:
        event_type = event.get("type", "")
        if event_type == "property_unavailable":
            result.notifications.append({"kind": "property_unavailable"})
        elif event_type == "new_property":
            result.notifications.append({"kind": "action_needed", "reason": "new_property_pending_send"})
        elif event_type == "call_requested":
            result.notifications.append({"kind": "action_needed", "reason": "call_requested"})
        elif event_type == "needs_user_input":
            reason = event.get("reason", "unknown")
            result.notifications.append({"kind": "action_needed", "reason": f"needs_user_input:{reason}"})

    # Check if row is complete
    required = ["total sf", "ops ex /sf", "drive ins", "docks", "ceiling ht", "power"]
    complete_count = 0
    for req in required:
        idx = header_lower.get(req)
        if idx is not None and idx < len(result.sheet_after) and result.sheet_after[idx]:
            complete_count += 1

    if complete_count == len(required):
        result.notifications.append({"kind": "row_completed"})

    # Validate against expected
    expected_updates = conversation.get("expected_updates", [])
    expected_events = conversation.get("expected_events", [])
    forbidden_updates = conversation.get("forbidden_updates", [])

    # Check expected updates
    actual_cols = {u.get("column", "").lower(): u.get("value") for u in result.ai_updates}
    for exp in expected_updates:
        col = exp.get("column", "")
        val = exp.get("value", "")
        if col.lower() not in actual_cols:
            result.issues.append(f"Missing expected update: {col}")
        elif actual_cols[col.lower()] != val:
            result.warnings.append(f"Value mismatch for {col}: expected '{val}', got '{actual_cols[col.lower()]}'")

    # Check forbidden updates
    for forbidden in forbidden_updates:
        if forbidden.lower() in actual_cols:
            result.issues.append(f"FORBIDDEN update detected: {forbidden}")

    # Check expected events
    actual_events = {e.get("type") for e in result.ai_events}
    for exp_event in expected_events:
        if exp_event not in actual_events:
            result.issues.append(f"Missing expected event: {exp_event}")

    # Determine pass/fail
    result.passed = len(result.issues) == 0

    return result


def display_result(result: E2ETestResult, header: List[str], verbose: bool = True):
    """Display a test result with full details."""
    print(f"\n{'='*70}")
    print(f"E2E Test: {result.property_address}")
    print(f"{'='*70}")

    # Show AI response summary
    print(f"\n📤 AI Response:")
    print(f"   Updates: {len(result.ai_updates)}")
    for u in result.ai_updates:
        print(f"      - {u.get('column')}: {u.get('value')}")
    print(f"   Events: {[e.get('type') for e in result.ai_events]}")

    # Show sheet state (before → after)
    print(f"\n📊 SHEET STATE:")
    print(f"   {'─'*66}")
    print(f"   │ {'Column':<25} │ {'Before':<15} │ {'After':<15} │")
    print(f"   {'─'*66}")

    for i, col in enumerate(header):
        if not col:
            continue
        before = result.sheet_before[i] if i < len(result.sheet_before) else ""
        after = result.sheet_after[i] if i < len(result.sheet_after) else ""

        # Only show rows with data or changes
        if before or after:
            changed = " ✏️" if before != after else ""
            print(f"   │ {col:<25} │ {str(before)[:15]:<15} │ {str(after)[:15]:<15} │{changed}")

    print(f"   {'─'*66}")

    # Show required fields status
    required = ["Total SF", "Ops Ex /SF", "Drive Ins", "Docks", "Ceiling Ht", "Power"]
    header_lower = {h.lower().strip() if h else "": i for i, h in enumerate(header)}
    complete = []
    missing = []
    for req in required:
        idx = header_lower.get(req.lower())
        if idx is not None and idx < len(result.sheet_after) and result.sheet_after[idx]:
            complete.append(req)
        else:
            missing.append(req)

    if complete:
        print(f"   ✅ Complete: {', '.join(complete)}")
    if missing:
        print(f"   ❌ Missing: {', '.join(missing)}")

    # Show response email
    print(f"\n📧 RESPONSE EMAIL:")
    if result.ai_response:
        print(f"   {'─'*56}")
        for line in result.ai_response.split('\n'):
            print(f"   │ {line}")
        print(f"   │ ")
        print(f"   │ Best,")
        print(f"   │ Jill")
        print(f"   {'─'*56}")
    else:
        print(f"   (none - escalated to user)")

    # Show notifications
    print(f"\n🔔 NOTIFICATIONS:")
    for n in result.notifications:
        if n["kind"] == "sheet_update":
            print(f"   - sheet_update: {n.get('column')} = {n.get('value')}")
        elif n["kind"] == "row_completed":
            print(f"   - row_completed ✅")
        else:
            reason = f" ({n.get('reason')})" if n.get('reason') else ""
            print(f"   - {n['kind']}{reason}")

    # Show result
    print(f"\n{'✅ PASS' if result.passed else '❌ FAIL'}")
    for issue in result.issues:
        print(f"   ❌ {issue}")
    for warning in result.warnings:
        print(f"   ⚠️ {warning}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description="E2E Integration Tests")
    parser.add_argument("-p", "--property", help="Run test for specific property")
    parser.add_argument("--list", action="store_true", help="List available conversations")
    parser.add_argument("-q", "--quiet", action="store_true", help="Minimal output")
    parser.add_argument("--edge-cases", action="store_true", help="Run edge case tests")
    parser.add_argument("--all", action="store_true", help="Run all tests including edge cases")
    parser.add_argument("--save", action="store_true", help="Save results to tests/results/")
    parser.add_argument("--list-runs", action="store_true", help="List previous test runs")
    parser.add_argument("--compare", nargs=2, metavar=("RUN1", "RUN2"), help="Compare two runs")
    parser.add_argument("--scrub", help="Path to custom Scrub Excel file")
    parser.add_argument("--generated", help="Run generated conversations from category (response_type, event, edge_case, format, or 'all')")
    args = parser.parse_args()

    # List previous runs
    if args.list_runs:
        runs = list_runs()
        if not runs:
            print("\nNo previous test runs found.")
            print("Run with --save to create result files.")
            return

        print("\nPrevious test runs:")
        print(f"{'─'*70}")
        for run in runs:
            status = f"{run.get('tests_passed', 0)}/{run.get('tests_run', 0)} passed"
            input_file = run.get('input_file', {}).get('filename', 'unknown')
            print(f"  {run['run_name']}")
            print(f"    Input: {input_file} | {status}")
            print(f"    Created: {run.get('created_at', 'unknown')}")
        return

    # Compare two runs
    if args.compare:
        comparison = compare_runs(args.compare[0], args.compare[1])
        if "error" in comparison:
            print(f"Error: {comparison['error']}")
            return

        print(f"\nComparing {args.compare[0]} vs {args.compare[1]}")
        print(f"{'─'*70}")

        if comparison.get("input_file_changed"):
            print("  ⚠️  Input Excel file has changed between runs")

        changes = comparison.get("changes", [])
        if not changes:
            print("  ✅ No differences found")
        else:
            for change in changes:
                prop = change.get("property")
                change_type = change.get("change")
                if change_type == "added_in_run2":
                    print(f"  + {prop} (new in {args.compare[1]})")
                elif change_type == "removed_in_run2":
                    print(f"  - {prop} (removed in {args.compare[1]})")
                elif change_type == "status_changed":
                    old = "✅" if change.get("run1_passed") else "❌"
                    new = "✅" if change.get("run2_passed") else "❌"
                    print(f"  Δ {prop}: {old} → {new}")
                elif change_type == "output_changed":
                    print(f"  Δ {prop}: output changed")
        return

    # List conversations
    if args.list:
        print("\nAvailable conversation files:")
        conv_dir = get_conversations_dir()
        if conv_dir.exists():
            print("\n  Main scenarios:")
            for f in sorted(conv_dir.glob("*.json")):
                print(f"    - {f.stem}")

            edge_dir = conv_dir / "edge_cases"
            if edge_dir.exists():
                print("\n  Edge cases:")
                for f in sorted(edge_dir.glob("*.json")):
                    print(f"    - edge_cases/{f.stem}")
        else:
            print(f"  (No conversations directory at {conv_dir})")
            print(f"  Create it and add JSON conversation files.")
        return

    # Load scrub file
    scrub_filepath = args.scrub or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "Scrub Augusta GA.xlsx"
    )
    print(f"Loading Scrub file: {os.path.basename(scrub_filepath)}...")
    try:
        properties = load_scrub_file(scrub_filepath)
        print(f"Loaded {len(properties)} properties from Scrub file")
    except Exception as e:
        print(f"Failed to load Scrub file: {e}")
        sys.exit(1)

    # Setup results saving if requested
    run_dir = None
    manifest = None
    if args.save:
        run_dir = create_run_directory()
        manifest = create_manifest(run_dir, scrub_filepath, properties)
        print(f"Results will be saved to: {run_dir}")

    # Ensure conversations directory exists
    conv_dir = get_conversations_dir()
    if not conv_dir.exists():
        print(f"\nCreating conversations directory: {conv_dir}")
        conv_dir.mkdir(parents=True)
        print("Add JSON conversation files to define test scenarios.")
        print("Example: conversations/699_industrial_park_dr.json")
        sys.exit(0)

    results = []

    # Run main property tests (unless --edge-cases only)
    if not args.edge_cases:
        # Filter to specific property if requested
        test_properties = properties
        if args.property:
            matching = {k: v for k, v in properties.items() if args.property.lower() in k.lower()}
            if not matching:
                print(f"No properties matching '{args.property}'")
                print("Available properties:")
                for addr in properties:
                    print(f"  - {addr}")
                sys.exit(1)
            test_properties = matching

        # Run tests
        for address, prop_data in test_properties.items():
            conversation = load_conversation(address)
            if not conversation:
                if not args.quiet:
                    print(f"\n⏭️  Skipping {address} (no conversation file)")
                continue

            if not args.quiet:
                print(f"\n🧪 Testing: {address}")

            result = run_e2e_test(address, prop_data, conversation)
            results.append(result)

            if not args.quiet:
                display_result(result, prop_data["header"])

            # Save result if requested
            if run_dir:
                save_result(run_dir, result, prop_data, conversation, prop_data["header"])

    # Run edge case tests
    if args.edge_cases or args.all:
        edge_cases = load_all_edge_case_conversations()
        if edge_cases:
            print(f"\n{'='*70}")
            print("EDGE CASE TESTS")
            print(f"{'='*70}")

            for filename, conv in edge_cases:
                prop_address = conv.get("property", "")
                if prop_address not in properties:
                    if not args.quiet:
                        print(f"\n⏭️  Skipping edge case '{filename}' (property '{prop_address}' not in Scrub)")
                    continue

                prop_data = properties[prop_address]

                if not args.quiet:
                    print(f"\n🧪 Edge case: {filename}")
                    print(f"   {conv.get('description', '')}")

                result = run_e2e_test(f"[EDGE] {filename}", prop_data, conv)
                results.append(result)

                if not args.quiet:
                    display_result(result, prop_data["header"])

                # Save result if requested
                if run_dir:
                    save_result(run_dir, result, prop_data, conv, prop_data["header"])

    # Run generated conversation tests
    if args.generated:
        generated = load_generated_conversations(args.generated)
        if generated:
            print(f"\n{'='*70}")
            print(f"GENERATED TESTS ({args.generated})")
            print(f"{'='*70}")

            for label, conv in generated:
                prop_address = conv.get("property", "")
                if prop_address not in properties:
                    if not args.quiet:
                        print(f"\n⏭️  Skipping '{label}' (property '{prop_address}' not in Scrub)")
                    continue

                prop_data = properties[prop_address]

                if not args.quiet:
                    print(f"\n🧪 Generated: {label}")
                    print(f"   {conv.get('description', '')}")

                result = run_e2e_test(label, prop_data, conv)
                results.append(result)

                if not args.quiet:
                    display_result(result, prop_data["header"])

                # Save result if requested
                if run_dir:
                    save_result(run_dir, result, prop_data, conv, prop_data["header"])
        else:
            print(f"\nNo generated conversations found for category: {args.generated}")
            print("Run: python3 tests/conversation_generator.py --generate-all")

    # Summary
    if results:
        passed = sum(1 for r in results if r.passed)
        print(f"\n{'='*70}")
        print("E2E TEST SUMMARY")
        print(f"{'='*70}")
        print(f"Total: {len(results)} | Passed: {passed} | Failed: {len(results) - passed}")
        print(f"Pass Rate: {passed/len(results)*100:.1f}%")

        if passed < len(results):
            print("\nFailed tests:")
            for r in results:
                if not r.passed:
                    print(f"  ❌ {r.property_address}")
                    for issue in r.issues:
                        print(f"      - {issue}")

        # Save summary if requested
        if run_dir and manifest:
            summary = save_summary(run_dir, results, manifest)
            print(f"\n📁 Results saved to: {run_dir}")
            print(f"   - manifest.json (run metadata)")
            print(f"   - summary.json (campaign summary)")
            print(f"   - {len(results)} individual result files")
    else:
        print("\nNo tests run. Create conversation files in tests/conversations/")
        print("Example filename: 699_industrial_park_dr.json")


if __name__ == "__main__":
    main()
